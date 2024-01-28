import openpyxl

wb = openpyxl.load_workbook("planilla.xlsx")

hoja = wb.active

wb.create_sheet("Hoja 3")

hoja3 = wb["Hoja 3"]
hoja3.title = "Nuevo titulo"

# print(
#     hoja.max_row,
#     hoja.max_column,
# )

celda = hoja["A1"]
celda.value = "Nombre completo"
# print(celda.value)

celda2 = hoja.cell(row=2, column=1)
# print(
#     celda2.value,
#     celda2.row,
#     celda2.column,
#     celda2.coordinate
# )

for fila in range(1, hoja.max_row + 1):
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila, column=columna)
        # print(fila, columna, celda.value)

columna = hoja["A"]
fila = hoja["1"]
# print(columna)
# print(fila)

hoja.append([1, 2, 3])
print(hoja.rows)
hoja.delete_rows(1, 1)
wb.save("nuevo_excel.xlsx")
